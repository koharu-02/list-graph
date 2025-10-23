import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
import plotly.express as px

st.set_page_config(layout="wide")
st.title("要素作業リスト生成＆工程編成検討ツール")

uploaded_file = st.file_uploader("Excelファイルをアップロードしてください", type=["xlsx"])

def to_number(val):
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0

if uploaded_file:
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    output = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        process = ws["D2"].value if ws["D2"].value else ""
        last_foot_pos = ""

        for i in range(6, 53):
            task = ws[f"B{i}"].value if ws[f"B{i}"].value else ""
            foot_pos = ws[f"Q{i}"].value if ws[f"Q{i}"].value else ""
            walk_time = ws[f"Q{i+1}"].value if ws[f"Q{i+1}"].value else ""
            task_time = ws[f"AB{i}"].value if ws[f"AB{i}"].value else ""

            if not task:
                continue

            if foot_pos:
                last_foot_pos = foot_pos
            else:
                foot_pos = last_foot_pos

            if to_number(walk_time) > 0:
                output.append({"工程": process, "作業位置": "", "要素作業": "歩行", "時間": to_number(walk_time)})
                output.append({"工程": process, "作業位置": foot_pos, "要素作業": task,
                               "時間": max(to_number(task_time) - to_number(walk_time), 0)})
            else:
                output.append({"工程": process, "作業位置": foot_pos, "要素作業": task, "時間": to_number(task_time)})

    df_original = pd.DataFrame(output)

    # ID割り振り
    df = df_original.copy()
    ids = []
    current_id = 1
    n = len(df)
    for i in reversed(range(n)):
        if df.loc[i, "要素作業"] == "歩行":
            if i + 1 < n and df.loc[i, "工程"] == df.loc[i + 1, "工程"]:
                ids.append(None)
            else:
                ids.append(current_id)
                current_id += 1
        else:
            ids.append(current_id)
            current_id += 1
    ids = ids[::-1]
    for i in range(n):
        if ids[i] is None:
            ids[i] = ids[i + 1]
    df["ID"] = ids

    st.subheader("元データ（ID割り振り済）")
    st.dataframe(df)

    df["ラベル"] = "ID:" + df["ID"].astype(str) + " | " + df["作業位置"].fillna("なし") + " | " + df["要素作業"] + " | " + df["時間"].astype(str) + "秒"
    df["色分けカテゴリ"] = df["作業位置"].where(df["作業位置"].notna(), df["要素作業"])

    fig = px.bar(
        df,
        x="工程",
        y="時間",
        color="色分けカテゴリ",
        text="ラベル",
        hover_data=["ID", "作業位置", "要素作業", "時間"],
        title="工程別作業時間（作業位置または要素作業ごとに積み上げ）"
    )
    fig.update_traces(marker=dict(line=dict(color="black", width=1)))
    fig.update_layout(
        barmode="stack",
        xaxis_title="工程",
        yaxis_title="時間",
        showlegend=False,
        height=600,
        margin=dict(l=40, r=40, t=60, b=40)
    )
    st.plotly_chart(fig, use_container_width=True)

    st.subheader("IDごとに移動先工程を指定（直接入力）")
    id_input = st.text_input("移動したいIDをカンマ区切りで入力してください（例: 1,2,5）")

    try:
        selected_ids = [int(x.strip()) for x in id_input.split(",") if x.strip()]
    except ValueError:
        st.error("IDは数値で入力してください。")
        selected_ids = []

    move_targets = {}
    for id_ in selected_ids:
        if id_ in df["ID"].values:
            current_process = df.loc[df["ID"] == id_, "工程"].values[0]
            move_targets[id_] = st.selectbox(
                f"ID:{id_}（現在：{current_process}）の移動先工程",
                options=[x for x in sorted(df["工程"].unique()) if x != current_process],
                key=f"move_{id_}"
            )
        else:
            st.warning(f"ID:{id_} はデータに存在しません。")

    if st.button("✅ 一括移動実行"):
        for id_, to_process in move_targets.items():
            df.loc[df["ID"] == id_, "工程"] = to_process

        st.success(f"{len(move_targets)} 件のIDの移動を実行しました。")

        df["ラベル"] = "ID:" + df["ID"].astype(str) + " | " + df["作業位置"].fillna("なし") + " | " + df["要素作業"] + " | " + df["時間"].astype(str) + "秒"
        df["色分けカテゴリ"] = df["作業位置"].where(df["作業位置"].notna(), df["要素作業"])

        fig_updated = px.bar(
            df,
            x="工程",
            y="時間",
            color="色分けカテゴリ",
            text="ラベル",
            hover_data=["ID", "作業位置", "要素作業", "時間"],
            title="更新後の工程別作業時間（作業位置または要素作業ごとに積み上げ）"
        )
        fig_updated.update_traces(marker=dict(line=dict(color="black", width=1)))
        fig_updated.update_layout(
            barmode="stack",
            xaxis_title="工程",
            yaxis_title="時間",
            showlegend=False,
            height=600,
            margin=dict(l=40, r=40, t=60, b=40)
        )
        st.plotly_chart(fig_updated, use_container_width=True)

    # Excelファイルをメモリ上に作成（元データと更新後データの両方を含む）
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_original.to_excel(writer, sheet_name="元データ", index=False)
        df.drop(columns=["色分けカテゴリ"]).to_excel(writer, sheet_name="更新後データ", index=False)
    buffer.seek(0)

    st.download_button("📥 Excelファイルをダウンロード（元データ＋更新後データ）", buffer, file_name="process_plan_combined.xlsx")
